import csv
import openpyxl
from openpyxl import load_workbook

#Create a new xlsx file
wb = openpyxl.Workbook()
sheet = wb.create_sheet("Values")

#Variable declaration
yes_no = 0
new_item = 0
cell = 0
counter = 0
counter2 = 0
longitude = 0
latitude = 0
allvalues = []
counter3 = 0
turns = 0
char1 = ","
char2 = ":"

while yes_no == 0:
    
    counter3 = 0
    error = True
    
    #Tries to open the selected file. In case the file is incorrect, it give possibles reason for the error.
    while error == True:
    
        try:
            file = input("Enter file name: ")
            with open(file) as f:
                
                #Searchs in the file for the the incial and final date, and the brand.
                lis = [line.split() for line in f]
                brand_line = (lis[2])
        
                brand = brand_line[0]
                brand = brand[brand.find(char1)+1 : brand.find(char2)]
        
                inicial_date = brand_line[1]
                inicial_date = inicial_date[1:]
        
                final_date = brand_line[3]
                final_date = final_date[:-1]
                
            error = False
            
        except:
            print("\n" + "File not found! Please check the following: " + "\n" + "- If the file is in the same folder" + "\n" + "- If the '.csv' is inserted at the end of the file name" + "\n" "- If the file name was introduced correctly")
        
        
    with open(file) as f:    
        reader = csv.reader(f) 
        
        for row in reader:
            
            #This block changes the row number, in order to eliminate unwanted rows. 
            #Depending on the number of files that are being read, its the amount of rows that are eliminated.
            counter3 = counter3 + 1
            if counter3 == 4:
                counter = counter - 2
                counter2 = counter2 - 2
            if counter3 == 4 and turns >= 1:
                counter = counter - 1
                counter2 = counter2 - 1
                
            for item in row:
                new_item = item
                
                #This block corrects all strings that are wrong due to invalid characters.
                if item == "TucumÃ¡n":
                    new_item = "Tucuman"
                if item == "CÃ³rdoba":
                    new_item = "Cordoba"
                if item == "Entre RÃ­os":
                    new_item = "Entre Rios"
                if item == "NeuquÃ©n":
                    new_item = "Neuquen"
                if item == "Ciudad AutÃ³noma de Buenos Aires":
                    new_item = "Capital"
                if item == "RÃ­o Negro":
                    new_item = "Rio Negro"
                if item == "MisiÃ³nes":
                    new_item = "Misiones"
        
                if new_item != "":
                    
                    try:
                        #it appends to the new excel file the number of searches in google in the correct row and column
                        new_item = int(new_item)
                        counter2 = counter2 + 1
                        sheet["F" + str(counter)] = new_item
                        
                    except:
                        #it appends to the new excel file the province name in the correct row and column
                        counter = counter + 1
                        sheet["D" + str(counter)] = new_item
                        
                        #Opens an excel file with the information of the latitudee and longitude of the province
                        wb2 = load_workbook("Latitude and longitude of each province.xlsx")
                        ws = wb2.active
                        allvalues = []
                        
                        #Searchs for the row in which the province information is
                        for row2 in ws.rows:
                            if row2[1].value == new_item:
                                for cell in row2:
                                    allvalues.append(cell.value)
                        
                        try:
                            #Gets the latitudee and longitudee of the province
                            longitude = allvalues[2]
                            latitude = allvalues[3]
                            
                            #Appends the rest of the values into the new excel file
                            sheet["G" + str(counter)] = longitude
                            sheet["H" + str(counter)] = latitude
                            sheet["E" + str(counter)] = brand
                            sheet["A" + str(counter)] = inicial_date
                            sheet["B" + str(counter)] = final_date
                            sheet["C" + str(counter)] = "Argentina"
                        
                        #I know this is an awful algorithm, but it works for this type of info in particular.
                        except:
                            continue
    
    print("\n" + "Analysis complete!")
    
    #Asks the user if it wants to analyze another file
    while error == False:
        new_file = input("Do you want to read another file(y/n): ")
        if new_file == "y" or new_file == "n":
            error = True
        else:
            print("\n" + "Invalid Character! Please enter 'y' for yes or 'n' for no" + "\n" + "Try again")
    if new_file == "n":
        yes_no = 1
    
    #This varibles helps Python understand if its the first file read or no.
    turns = turns + 1

#Insert column titles
sheet["G1"] = "Longitude"
sheet["H1"] = "Latitude"
sheet["E1"] = "Brand"
sheet["A1"] = "Inicial Date"
sheet["B1"] = "Final Date"
sheet["C1"] = "Country"
sheet["D1"] = "Province"
sheet["F1"] = "Searches"

#Saves the new file
while error == True:
    try:     
        wb.save("GOOGLE TRENDS.xlsx")
        error = False
    except:
        print("/n" + "GOOGLE TRENDS.xlsx is open. Please close it and press enter")
        enter= input()