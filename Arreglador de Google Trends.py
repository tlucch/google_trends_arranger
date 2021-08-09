import csv
import openpyxl
from openpyxl import load_workbook

wb = openpyxl.Workbook()
hoja = wb.create_sheet("hoja")

sino = 0
new_item = 0
celda = 0
contador = 0
contador2 = 0
longitud = 0
latitud = 0
allvalues = []
contador3 = 0
vueltas = 0
char1 = ","
char2 = ":"

while sino == 0:
    
    contador3 = 0
    error = True
    
    while error == True:
        try:
            Archivo = input("Introduzca el nombre del archivo: ")
            with open(Archivo) as f:
                lis = [line.split() for line in f]
                lineamarca = (lis[2])
        
                marca = lineamarca[0]
                marca = marca[marca.find(char1)+1 : marca.find(char2)]
        
                fechai = lineamarca[1]
                fechai = fechai[1:]
        
                fechaf = lineamarca[3]
                fechaf = fechaf[:-1]
            error = False
        except:
            print("Archivo no encontrado! Chequear lo siguiente: " + "\n" + "- si el archivo esta en la carpeta" + "\n" + "- si inserto el .csv al final del nombre del archivo" + "\n" "- si el nombre del archivo esta escrito correctamente")
            
    with open(Archivo) as f:               
        reader = csv.reader(f)     
        
        for row in reader:
            contador3 = contador3 + 1
            if contador3 == 4:
                contador = contador - 2
                contador2 = contador2 - 2
            if contador3 == 4 and vueltas >= 1:
                contador = contador - 1
                contador2 = contador2 - 1
            for item in row:
                new_item = item
                if item == "TucumÃ¡n":
                    new_item = "Tucuman"
                if item == "CÃ³rdoba":
                    new_item = "Cordoba"
                if item == "Entre RÃ­os":
                    new_item = "Entre Rios"
                if item == "NeuquÃ©n":
                    new_item = "Neuquen"
                if item == "Ciudad AutÃ³noma de Buenos Aires":
                    new_item = "Capital"
                if item == "RÃ­o Negro":
                    new_item = "Rio Negro"
                if item == "MisiÃ³nes":
                    new_item = "Misiones"
        
                if new_item != "":
                    
                    try:
                        new_item = int(new_item)
                        contador2 = contador2 + 1
                        hoja["F" + str(contador)] = new_item
                        
                    except:
                        contador = contador + 1
                        hoja["D" + str(contador)] = new_item
                        
                        wb2 = load_workbook("latitud y longitud por provincia.xlsx")
                        ws = wb2.active
                        allvalues = []
                        
                        for row2 in ws.rows:
                            if row2[1].value == new_item:
                                for cell in row2:
                                    allvalues.append(cell.value)
                                    
                        try:
                            longitud = allvalues[2]
                            latitud = allvalues[3]
    
                            hoja["G" + str(contador)] = longitud
                            hoja["H" + str(contador)] = latitud
                            hoja["E" + str(contador)] = marca
                            hoja["A" + str(contador)] = fechai
                            hoja["B" + str(contador)] = fechaf
                            hoja["C" + str(contador)] = "Argentina"
                            
                        except:
                            continue
    
    while error == False:
        otroarchivo = input("Queres leer otro archivo(y/n): ")
        if otroarchivo == "y" or otroarchivo == "n":
            error = True
        else:
            print("Caracter invalido! Por favor introduzca 'y' para si o 'n' para no" + "\n" + "Intente nuevamente")
            
    if otroarchivo == "n":
        sino = 1
    
    vueltas = vueltas + 1

hoja["G1"] = "Longitud"
hoja["H1"] = "Latitud"
hoja["E1"] = "Marca"
hoja["A1"] = "Fecha Inicio"
hoja["B1"] = "Fecha Finalizacion"
hoja["C1"] = "Pais"
hoja["D1"] = "Provincia"
hoja["F1"] = "Busquedas"

while error == True:
    try:     
        wb.save("GOOGLE TRENDS.xlsx")
        error = False
    except:
        print("Archivo GOOGLE TRENDS.xlsx abierto. Por favor cierrelo y presione enter")
        enter= input()
